# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import openpyxl
import io
import zipfile

# === アプリのタイトル ===
st.title("📂 Excel列分解＆CSV作成ツール")
st.write("基準となる列（店舗名など）と、数式列をペアにしてCSV化します。")

# === サイドバー：設定エリア ===
st.sidebar.header("⚙️ 設定")

# 1. 基準列の設定
st.sidebar.subheader("1. 基準列（店舗名）")
anchor_col = st.sidebar.text_input("固定して使う列 (例: A)", value="A")

# 2. 行の削除設定
st.sidebar.subheader("2. 行の削除")
skip_rows = st.sidebar.number_input("最初に削除する行数", min_value=0, value=2)

# 3. 重複削除の設定
st.sidebar.subheader("3. データの整理")
remove_dup = st.sidebar.checkbox("重複した行を自動で削除する", value=True)
st.sidebar.caption("※同じ店舗名・同じ数値の行がある場合、最初の行を残します。")

# 4. 列の除外設定
st.sidebar.subheader("4. 列の除外設定")
ignore_col_start = st.sidebar.text_input("除外したい開始列 (例: B)", value="")
ignore_col_end = st.sidebar.text_input("除外したい終了列 (例: G)", value="")

# === メインエリア ===
uploaded_file = st.file_uploader("Excelファイルをアップロード", type=['xlsx'])

if uploaded_file:
    try:
        file_bytes = uploaded_file.getvalue()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
        ws = wb.active
        
        start_row = skip_rows + 1
        max_check = min(start_row + 10, ws.max_row)
        
        try:
            anchor_idx = openpyxl.utils.column_index_from_string(anchor_col) - 1
        except:
            st.error("基準列の指定が間違っています。")
            st.stop()

        # ---------------------------------------------------------
        # 元データの読み込みと店舗リストの基準作成
        # ---------------------------------------------------------
        df_full = pd.read_excel(
            io.BytesIO(file_bytes), 
            header=None, 
            skiprows=skip_rows, 
            engine='openpyxl'
        )
        # 基準となる店舗名リスト（前後の空白除去）
        original_stores = df_full.iloc[:, anchor_idx].astype(str).str.strip().tolist()
        st.info(f"📊 元データの店舗数: {len(original_stores)} 件")

        # --- 除外設定 ---
        ignore_indices = []
        if ignore_col_start and ignore_col_end:
            try:
                start_ignore = openpyxl.utils.column_index_from_string(ignore_col_start)
                end_ignore = openpyxl.utils.column_index_from_string(ignore_col_end)
                ignore_indices = list(range(start_ignore - 1, end_ignore))
            except:
                st.error("除外列の指定が間違っています。")

        # --- 数式列の検出 ---
        formula_candidates = []
        for col_idx_1based in range(1, ws.max_column + 1): 
            col_idx_0based = col_idx_1based - 1
            if col_idx_0based == anchor_idx or col_idx_0based in ignore_indices:
                continue

            is_formula = False
            for r in range(start_row, max_check + 1):
                cell = ws.cell(row=r, column=col_idx_1based)
                if cell.data_type == 'f' or (str(cell.value).startswith('=')):
                    is_formula = True
                    break
            
            if is_formula:
                col_letter = openpyxl.utils.get_column_letter(col_idx_1based)
                formula_candidates.append({"idx": col_idx_0based, "name": col_letter})

        if not formula_candidates:
            st.warning("⚠️ 数式列が見つかりませんでした。")
        else:
            st.success(f"✅ {len(formula_candidates)} 個の数式列が見つかりました！")
            
            st.subheader("🛠️ 出力する列を選択")
            cols = st.columns(4)
            selected_indices = []
            for i, candidate in enumerate(formula_candidates):
                with cols[i % 4]:
                    if st.checkbox(f"{candidate['name']} 列", value=True, key=candidate['idx']):
                        selected_indices.append(candidate['idx'])

            # --- CSV作成 ---
            st.markdown("---")
            if st.button("🚀 選択した列のCSVを作成"):
                if not selected_indices:
                    st.error("列が選択されていません。")
                else:
                    with st.spinner('CSVを作成中...'):
                        zip_buffer = io.BytesIO()
                        validation_errors = []

                        with zipfile.ZipFile(zip_buffer, 'w') as myzip:
                            for target_idx in selected_indices:
                                col_letter = openpyxl.utils.get_column_letter(target_idx + 1)
                                cell_value_row2 = ws.cell(row=2, column=target_idx + 1).value
                                suffix = f"_{cell_value_row2}" if cell_value_row2 is not None else ""
                                filename = f"output_column_{col_letter}{suffix}.csv"
                                
                                # 抽出と加工
                                output_df = df_full.iloc[:, [anchor_idx, target_idx]].copy()
                                output_df.iloc[:, 0] = output_df.iloc[:, 0].astype(str).str.strip()
                                
                                if remove_dup:
                                    output_df = output_df.drop_duplicates(keep='first')

                                # ---------------------------------------------------------
                                # 順番と件数のチェック
                                # ---------------------------------------------------------
                                current_stores = output_df.iloc[:, 0].tolist()
                                if current_stores != original_stores:
                                    diff_count = len(original_stores) - len(current_stores)
                                    msg = f"【{col_letter}列】 元データと一致しません。"
                                    if diff_count > 0:
                                        msg += f"（重複などで {diff_count} 件減少）"
                                    else:
                                        msg += f"（並び順が変更されています）"
                                    validation_errors.append(msg)

                                # CSV書き出し
                                csv_data = output_df.to_csv(header=False, index=False, encoding='utf-8-sig')
                                myzip.writestr(filename, csv_data)
                        
                        # チェック結果の表示
                        if not validation_errors:
                            st.success("✅ 全てのファイルが元データの順番・件数通りに作成されました！")
                        else:
                            with st.expander("⚠️ データに変更がありました（重複削除などの結果）"):
                                for err in validation_errors:
                                    st.write(err)

                        st.download_button(
                            label="📥 ZIPファイルをダウンロード",
                            data=zip_buffer.getvalue(),
                            file_name="処理結果.zip",
                            mime="application/zip"
                        )

    except Exception as e:
        st.error(f"エラーが発生しました: {e}")
